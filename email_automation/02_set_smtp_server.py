import os
import smtplib
from email.mime.text import MIMEText
from openpyxl.reader.excel import load_workbook


class EmailSender:
    email_addr = None
    password = None
    smtp_server_map = {
        'gmail.com': 'smtp.gmail.com',
        'naver.com': 'smtp.naver.com'
    }
    smtp_server = None

    def __init__(self, email_addr, password):
        print('생성자')
        self.email_addr = email_addr
        self.password = password
        self.smtp_server = self.smtp_server_map[email_addr.split('@')[1]]   # h01041303675@gmail.com
        print(self.smtp_server)

    def send_email(self, msg, from_addr, to_addr, subject):
        """
        :param msg: 보낼 메시지
        :param from_addr: 보내는 사람
        :param to_addr: 받는 사람
        :return:
        """
        # with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            msg = MIMEText(msg)     # 네이버 일 때만 필요 - 시작
            msg['From'] = from_addr
            msg['To'] = to_addr
            msg['Subject'] = subject    # 네이버 일 때만 필요 - 끝
            print(msg.as_string())

            smtp.starttls()
            smtp.login(self.email_addr, self.password)
            # smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.encode('utf-8'))
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print('이메일 전송이 완료 되었습니다.')

    def send_all_emails(self, filename):
        print(f'{filename}에 있는 이메일과 내용을 이용해 메일을 보냅니다.')
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows():
            print(row[0], row[1].value)


if __name__ == '__main__':
    # es = EmailSender('h01041303675@gmail.com', os.getenv('MY_GMAIL_PASSWORD'))
    es = EmailSender('jyhuh8775@naver.com', os.getenv('MY_NAVER_PASSWORD'))
    # es.send_email('테스트 입니다. 2', from_addr='h01041303675@gmail.com', to_addr='h01041303675@gmail.com')
    # es.send_email(
    #     msg='테스트 입니다.\n네이버 이메일에서 보냄',
    #     from_addr='jyhuh8775@naver.com',
    #     to_addr='h01041303675@gmail.com',
    #     subject='이메일 전송 테스트3',
    # )

    es.send_all_emails('이메일리스트.xlsx')