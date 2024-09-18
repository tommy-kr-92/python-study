import os
from datetime import datetime

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment

# pd.set_option('display.max_column', None)


class ClassificationExcel:
    path = ''

    def __init__(self, order_xlsx_file, partner_info_xlsx_filename, path='result'):
        # 주문 목록
        df = pd.read_excel(order_xlsx_file, engine='openpyxl')
        df = df.rename(columns=df.iloc[1])  # 첫번째 줄 열 제목으로 설정
        '''
        0                                      NaN        NaN  ...         NaN         NaN
        1                                     주문번호       상품번호  ...          주소     주문시요구사항
        '''
        '''
                                   주문번호    상품번호  ...      주소  주문시요구사항
        0                   NaN     NaN  ...     NaN      NaN
        1                  주문번호    상품번호  ...      주소  주문시요구사항
        '''
        # print(df['상품명'])
        df = df.drop([df.index[0], df.index[1]])  # 0번째 index 삭제 - NaN 삭제
        # print(df.count())   # 개수 확인

        df = df.reset_index(drop=True)
        self.order_list = df
        self.path = path
        # print(df)

        # 파트너 목록
        df_partners = pd.read_excel(partner_info_xlsx_filename, engine='openpyxl')

        self.brands = df_partners['브랜드'].to_list()  # 브랜드명 리스트화
        self.partners = df_partners['업체명'].to_list()  # 업체명 리스트화

    def classify(self):

        for i, row in self.order_list.iterrows():  # iterrow()는 두가지를 가져와야 함
            brand_name = ''
            partner_name = ''

            for j in range(len(self.brands)):
                # print(self.brands[j])
                if self.brands[j] in row['상품명']:
                    # print(f'{self.brands[j]} 이(가) {j}번째에 포함되어 있습니다.')
                    brand_name = self.brands[j]
                    partner_name = self.partners[j]
                    break

            # print(f'{row["상품명"]}은 {brand_name} 브랜드 입니다. {j}번째')
            # print(f'업체명: {partner_name}')
            # print('------------------------------------')
            # print(row['상품명'])

            if partner_name != '':
                # 필터링
                df_filtered = self.order_list[self.order_list['상품명'].str.contains(brand_name)]
                # print(df_filtered)
                df_filtered.to_excel(f'{self.path}/[패스트몰] {partner_name}.xlsx', engine='openpyxl')
            else:
                print('없는 brand name:', brand_name, row['상품명'])

        # print(len(self.brands), self.brands)
        # print(len(self.partners), self.partners)

    def set_form(self, file_name):
        # df 에는 꾸미는 기능이 없음
        wb = load_workbook(file_name)
        ws = wb.active
        print('value:', ws['B1'].value)
        print('value:', ws['B2'].value)

        # 개수 세기
        row_cnt = ws.max_row - 1    # max_row 최대 몇 행이 있는지 알 수 있음
        print('cnt:', row_cnt)

        # 열 삽입
        ws.insert_rows(1)
        ws.insert_rows(1)

        now_day = datetime.now().strftime('%Y-%m-%d')

        # A1
        ws['A1'] = f'발송요청내역 [총 {row_cnt}건] {now_day}'
        ws['A1'].font = Font(size=11, bold=True)
        ws.merge_cells('A1:U1')
        ws['A1'].alignment = Alignment(horizontal='left')

        wb.save(file_name)

    def set_forms(self):
        file_list = os.listdir(self.path)
        print(file_list)
        for file_name in file_list:
            file_name = f'{self.path}/{file_name}'
            self.set_form(file_name)


if __name__ == '__main__':
    ce = ClassificationExcel('주문목록20221112.xlsx', '파트너목록.xlsx', '20240918')
    # ce.classify()
    ce.set_forms()