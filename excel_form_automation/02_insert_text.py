from openpyxl import Workbook


class WeeklyWorkPlan:
    wb = None # workbook
    ws = None # worksheet
    start_date = "2024-09-17" # 시작일
    manager = "매니저 이름을 입력 해주세요" # 담당자 이름

    def __init__(self, start_date, manager,sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]
        self.start_date = start_date
        self.manager = manager
        self.set_title()

    def save(self, filename): # 엑셀 파일 저장 하는 함수
        self.wb.save(filename)
        print('엑셀 파일 생성 완료')

    def set_title(self): # 특정 cell 에 값을 넣는 함수
        ws = self.ws
        ws['B2'] = '담당자'
        ws['C2'] = self.manager
        ws['B3'] = '시작일'
        ws['C3'] = self.start_date

        # 제목
        ws['B5'] = '주간업무계획표'
        start_date = '2024-09-17'
        end_date = '2024-09-20'
        ws['B6'] = f'({start_date} ~ {end_date}'

        # 셀 병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

        print('타이틀 생성 완료')

        

if __name__ == '__main__':
    wwp = WeeklyWorkPlan('2024-09-20','허재영')
    wwp.save('주간업무계획표.xlsx')
