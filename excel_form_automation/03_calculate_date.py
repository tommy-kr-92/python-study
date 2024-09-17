from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook


class WeeklyWorkPlan:
    wb = None   # workbook
    ws = None   # worksheet
    start_date = "2024-09-17"   # 시작일
    manager = "매니저 이름을 입력 해주세요"     # 담당자 이름
    date_list = []
    days_of_week = []   # 요일

    def __init__(self, start_date, manager,sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]
        self.start_date = start_date
        self.manager = manager

        # 날짜 생성
        self.set_date()
        self.set_title()

    def save(self, filename):   # 엑셀 파일 저장 하는 함수
        self.wb.save(filename)
        print('엑셀 파일 생성 완료')

    def set_date(self, days=6):
        # start_date + 6
        end_date = datetime.strptime(self.start_date, '%Y-%m-%d') + timedelta(days=days)
        # 날짜 리스트에 넣기 - pandas 사용
        week = pd.date_range(start=self.start_date, end=end_date.strftime('%Y-%m-%d'))
        self.date_list = week.strftime('%Y-%m-%d').to_list()
        self.days_of_week = week.strftime('%A').to_list()

        # print('end_date:', end_date)
        # print('week:', week)
        print('date_list:', self.date_list)
        print('days_of_week:', self.days_of_week)


    def set_title(self):    # 특정 cell 에 값을 넣는 함수
        ws = self.ws
        ws['B2'] = '담당자'
        ws['C2'] = self.manager
        ws['B3'] = '시작일'
        ws['C3'] = self.start_date

        # 제목
        ws['B5'] = '주간업무계획표'
        start_date = self.date_list[0]
        end_date = self.date_list[-1]   # 리스트 중 제일 마지막
        ws['B6'] = f'({start_date} ~ {end_date})'

        # 셀 병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

        print('타이틀 생성 완료')



        

if __name__ == '__main__':
    wwp = WeeklyWorkPlan('2024-09-17','허재영')
    wwp.save('주간업무계획표.xlsx')
